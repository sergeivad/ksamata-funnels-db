import openpyxl

from db_source import FunnelRow
from findings import CLASS_TITLES, Finding
from report import SOURCES_SHEET, SUMMARY_SHEET, build_summary_rows, write_report


def finding(cls, funnel='f11'):
    return Finding(cls=cls, funnel=funnel, tag_type='reg', subject='S',
                   detail='D', evidence='E', first_seen='2026-05-02',
                   last_seen='2026-05-13', deals=3)


FUNNELS = [
    FunnelRow(funnel_id=11, num=11, front_code='f11', product_name='ДБО NR ВК', status='active'),
    FunnelRow(funnel_id=12, num=12, front_code='f12', product_name='ЖКТ NR ВК', status='active'),
]


def test_summary_counts_findings_per_funnel_and_class():
    rows = build_summary_rows([finding(1), finding(1), finding(4)], FUNNELS)
    header, first, second = rows[0], rows[1], rows[2]
    assert header[0] == 'Воронка'
    assert first[0] == 'f11'
    assert first[header.index('Класс 1')] == 2
    assert first[header.index('Класс 4')] == 1
    assert second[0] == 'f12'
    assert second[header.index('Класс 1')] == 0


def test_write_report_creates_summary_sources_and_one_sheet_per_class(tmp_path):
    out = tmp_path / 'report.xlsx'
    sources = [{'kind': 'выгрузка', 'name': 'deal_export_2026-05-02.csv', 'detail': '120 строк'}]
    write_report(str(out), [finding(1), finding(15)], FUNNELS, sources)

    wb = openpyxl.load_workbook(out)
    assert SUMMARY_SHEET in wb.sheetnames
    assert SOURCES_SHEET in wb.sheetnames
    # Лист заводится на каждый класс, даже пустой — чтобы «ноль находок»
    # был виден явно, а не выглядел как забытая проверка.
    for cls in CLASS_TITLES:
        assert f'Класс {cls}' in wb.sheetnames


def test_class_sheet_carries_decision_column_left_empty(tmp_path):
    out = tmp_path / 'report.xlsx'
    write_report(str(out), [finding(1)], FUNNELS, [])
    ws = openpyxl.load_workbook(out)['Класс 1']
    # _write_sheet: строка 1 — название листа, 2 — пустая, 3 — заголовки, 4+ — данные.
    header = [c.value for c in ws[3]]
    assert 'Решение' in header
    assert ws.cell(row=4, column=header.index('Решение') + 1).value is None


def test_class_sheet_title_row_names_the_class(tmp_path):
    out = tmp_path / 'report.xlsx'
    write_report(str(out), [], FUNNELS, [])
    ws = openpyxl.load_workbook(out)['Класс 16']
    assert CLASS_TITLES[16] in str(ws['A1'].value or ws['A2'].value or '')


def test_sources_sheet_lists_inputs(tmp_path):
    out = tmp_path / 'report.xlsx'
    sources = [
        {'kind': 'выгрузка', 'name': 'deal_export_2026-05-02.csv', 'detail': '120 строк'},
        {'kind': 'API', 'name': 'offer/get-offers', 'detail': '7679 предложений'},
    ]
    write_report(str(out), [], FUNNELS, sources)
    ws = openpyxl.load_workbook(out)[SOURCES_SHEET]
    values = [str(row[1].value) for row in ws.iter_rows(min_row=2) if row[1].value]
    assert 'deal_export_2026-05-02.csv' in values
    assert 'offer/get-offers' in values
