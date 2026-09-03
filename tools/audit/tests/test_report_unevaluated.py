#!/usr/bin/env python3
"""Отчёт обязан отличать «проверено, расхождений нет» от «не проверяли».

Пустой лист класса читается как «всё в порядке» — и это верно ровно тогда,
когда класс считался. Без реестра предложений GetCourse шесть классов не
считаются вовсе, и их пустота не значит ничего.

До этих тестов разницу держала только надпись в README. На надпись в
документации в этом пакете уже полагались однажды: сверка написания тега
этапа месяц отдавала ноль совпадений, и ноль читали как «расхождений нет»
(фаза 15). Ровно тот же симптом — молчание вместо ответа.
"""

import ast
import datetime
from collections import Counter
from pathlib import Path

import openpyxl
import pytest

import findings as findings_module
import run_audit
from export_source import Observation
from findings import group_observations, find_unresolved
from normalize import AUTOFUNNEL_TAG, parse_tagset

from db_source import FunnelRow
from findings import (
    CLASS_TITLES,
    REGISTRY_FILTERED_CLASSES,
    REGISTRY_ONLY_CLASSES,
    Finding,
    degraded_classes,
    unevaluated_classes,
)
from report import DEGRADED_NOTE, NOT_EVALUATED_NOTE, build_summary_rows, write_report

FUNNELS = [
    FunnelRow(funnel_id=11, num=11, front_code='f11', product_name='ДБО NR ВК',
              status='active', has_predspisok=True),
]


def finding(cls, funnel='f11'):
    return Finding(cls=cls, funnel=funnel, tag_type='reg', subject='S', detail='D',
                   evidence='E', first_seen='2026-05-02', last_seen='2026-05-13', deals=3)


def note_of(path, cls):
    """Пометка листа класса. Лежит в A2 — той строке, что и так была пустой:
    заголовки на 3-й, данные с 4-й, и сдвигать их нельзя."""
    return openpyxl.load_workbook(path)[f'Класс {cls}']['A2'].value


# --- какие классы без реестра слепнут -------------------------------------
#
# Множества ВЫВОДЯТСЯ ИЗ КОДА, а не переписываются литералом рядом с
# литералом. Утверждение `REGISTRY_ONLY_CLASSES == {9,...,17}` закрепляет
# константу, но не сверяет её с финдерами: новый registry-only класс 18,
# забытый в множестве, оставил бы такой тест зелёным, а лист печатал бы
# лживый `0` — ровно тот дефект, ради которого всё это и написано. Идиома
# та же, что у app/tests/migration-runners.test.ts: обойти код и спросить
# его самого.


def _finder_classes():
    """Финдер -> классы, которые он выдаёт.

    Литерала `cls=N` мало: find_unresolved кладёт номер в ПЕРЕМЕННУЮ
    (`cls, subject = 5, ...`), и сканер по ключевым аргументам его теряет —
    молча и с зелёным тестом. Собираем обе формы.
    """
    tree = ast.parse(Path(findings_module.__file__).read_text(encoding='utf-8'))
    out = {}
    for node in tree.body:
        if not isinstance(node, ast.FunctionDef) or not node.name.startswith('find_'):
            continue
        classes = set()
        for sub in ast.walk(node):
            if isinstance(sub, ast.Call):
                classes |= {kw.value.value for kw in sub.keywords
                            if kw.arg == 'cls' and isinstance(kw.value, ast.Constant)
                            and isinstance(kw.value.value, int)}
            if isinstance(sub, ast.Assign) and isinstance(sub.value, ast.Tuple):
                for target in sub.targets:
                    if not isinstance(target, ast.Tuple):
                        continue
                    for name, value in zip(target.elts, sub.value.elts):
                        if (isinstance(name, ast.Name) and name.id == 'cls'
                                and isinstance(value, ast.Constant)
                                and isinstance(value.value, int)):
                            classes.add(value.value)
        if classes:
            out[node.name] = classes
    return out


def _registry_args_by_finder():
    """Финдер -> что из реестра ему отдаёт collect_findings.

    Смотрим на РЕАЛЬНУЮ проводку в run_audit, а не на сигнатуры: важно не
    то, что функция умеет принять, а то, что ей отдают в прогоне. Развилка
    там ровно одна — финдеру дают либо сырой реестр (`offers`), либо
    производный от него фильтр (`registry_keys` / `registry_tags`).
    """
    tree = ast.parse(Path(run_audit.__file__).read_text(encoding='utf-8'))
    body = next(n for n in tree.body
                if isinstance(n, ast.FunctionDef) and n.name == 'collect_findings')
    return {call.func.attr: {a.id for a in call.args if isinstance(a, ast.Name)}
            for call in ast.walk(body)
            if isinstance(call, ast.Call) and isinstance(call.func, ast.Attribute)
            and call.func.attr.startswith('find_')}


def test_registry_only_classes_are_derived_from_the_finders_given_the_raw_registry():
    """Кому отдают сам реестр — тот без реестра не даёт ничего."""
    wiring = _registry_args_by_finder()
    classes = _finder_classes()
    derived = set()
    for name, args in wiring.items():
        if 'offers' in args:
            derived |= classes.get(name, set())
    assert derived == REGISTRY_ONLY_CLASSES


def test_registry_filtered_finders_are_derived_but_class_5_is_excluded_by_hand():
    """Кому отдают производный фильтр — тот считается, но с завышением.

    Вывести это множество ДО КОНЦА нельзя, и в этом вся ловушка:
    find_unresolved выдаёт два класса, 5 и 7, а фильтр стоит внутри ветки,
    которая ставит cls = 7. Поэтому здесь выводятся ФИНДЕРЫ, а разделение
    их классов закрепляет тест ниже — поведением, а не разбором текста.
    """
    wiring = _registry_args_by_finder()
    classes = _finder_classes()
    filtered = {name for name, args in wiring.items()
                if args & {'registry_keys', 'registry_tags'}}
    assert filtered == {'find_extra_axes', 'find_unresolved'}
    assert classes['find_unresolved'] == {5, 7}
    assert REGISTRY_FILTERED_CLASSES < set().union(*(classes[n] for n in filtered))
    assert REGISTRY_FILTERED_CLASSES == {2, 7}


def test_registry_keys_filters_class_7_and_leaves_class_5_alone():
    """Поведенческое доказательство, что 5 в множество не входит по праву.

    Класс 5 приходит из той же функции, но зависимости от реестра у него
    нет: пустой registry_keys не меняет по нему ровным счётом ничего.
    Ложная пометка «без реестра число завышено» легла бы на честный лист.
    """
    orphan = ('АВ Продукт: ЩЖ|АВ Подрядчик: НИМБ|АВ Канал: Яндекс|'
              'АВ Направление: РСЯ|' + AUTOFUNNEL_TAG)
    av = 'АВ Продукт: ДБО|АВ Подрядчик: NR|АВ Канал: ВК|АВ Направление: In Stream'

    def obs(raw, deal_id):
        return Observation(deal_id=deal_id, tags=parse_tagset(raw),
                           file_name='deal_export_2026-05-02_00-00-00.csv',
                           file_date=datetime.date(2026, 5, 2),
                           deal_created='2026-05-01 00:00:00')

    groups = group_observations([
        obs(av + '|АВ Этап: Оплата', '1'),              # класс 5: оплата без времени
        obs(orphan + '|АВ Этап: Регистрация', '2'),     # класс 7: ключа нет в базе
    ])
    index = {('ДБО', 'NR', 'ВК', 'In Stream', None): {11}}

    without = Counter(f.cls for f in find_unresolved(groups, index))
    with_registry = Counter(f.cls for f in find_unresolved(
        groups, index, registry_keys={('нет', 'такого', 'ключа', 'в', 'реестре')}))

    assert without == {5: 1, 7: 1}
    assert with_registry[5] == without[5]   # фильтр пятого не касается
    assert with_registry[7] == 0            # а седьмой гасит целиком


# --- признак «реестр не читали» — данные, а не флаг ------------------------

def test_empty_registry_marks_classes_regardless_of_why_it_is_empty():
    """Пустой реестр — и от `--no-api`, и от ответа API, из которого ничего
    не пришло. Класс слеп одинаково, значит и говорить надо одинаково."""
    assert unevaluated_classes([]) == REGISTRY_ONLY_CLASSES
    assert degraded_classes([]) == REGISTRY_FILTERED_CLASSES


def test_non_empty_registry_marks_nothing():
    offers = [object()]
    assert unevaluated_classes(offers) == frozenset()
    assert degraded_classes(offers) == frozenset()


# --- лист класса ----------------------------------------------------------

def test_unevaluated_class_sheet_says_so_in_plain_words(tmp_path):
    out = tmp_path / 'r.xlsx'
    write_report(str(out), [], FUNNELS, [], not_evaluated=REGISTRY_ONLY_CLASSES,
                 degraded=REGISTRY_FILTERED_CLASSES)
    for cls in REGISTRY_ONLY_CLASSES:
        assert note_of(out, cls) == NOT_EVALUATED_NOTE


def test_degraded_class_sheet_warns_about_inflated_counts(tmp_path):
    out = tmp_path / 'r.xlsx'
    write_report(str(out), [finding(2, funnel='—')], FUNNELS, [],
                 not_evaluated=REGISTRY_ONLY_CLASSES, degraded=REGISTRY_FILTERED_CLASSES)
    for cls in REGISTRY_FILTERED_CLASSES:
        assert note_of(out, cls) == DEGRADED_NOTE


def test_evaluated_class_sheet_carries_no_note(tmp_path):
    out = tmp_path / 'r.xlsx'
    write_report(str(out), [], FUNNELS, [], not_evaluated=REGISTRY_ONLY_CLASSES,
                 degraded=REGISTRY_FILTERED_CLASSES)
    for cls in set(CLASS_TITLES) - REGISTRY_ONLY_CLASSES - REGISTRY_FILTERED_CLASSES:
        assert note_of(out, cls) is None


def test_full_run_leaves_every_sheet_unmarked(tmp_path):
    """Прогон с реестром не должен пугать ни одним листом."""
    out = tmp_path / 'r.xlsx'
    write_report(str(out), [finding(1)], FUNNELS, [])
    for cls in CLASS_TITLES:
        assert note_of(out, cls) is None


def test_note_does_not_move_headers_or_data(tmp_path):
    """Пометка занимает уже существующую пустую строку 2. Заголовки обязаны
    остаться на 3-й, данные — с 4-й: на эти номера смотрит и test_report.py,
    и глаз человека, привыкшего к формату."""
    out = tmp_path / 'r.xlsx'
    write_report(str(out), [finding(17, funnel='f11')], FUNNELS, [],
                 not_evaluated=REGISTRY_ONLY_CLASSES)
    ws = openpyxl.load_workbook(out)['Класс 17']
    assert ws['A3'].value == 'Воронка'
    assert ws['A4'].value == 'f11'
    assert ws.freeze_panes == 'A4'


# --- сводка ---------------------------------------------------------------

def test_summary_header_marks_unevaluated_classes():
    rows = build_summary_rows([], FUNNELS, not_evaluated={17})
    header = rows[0]
    assert 'Класс 17 (не проверялся)' in header
    assert 'Класс 17' not in header
    assert 'Класс 16' in header  # остальные не тронуты


def test_summary_header_is_untouched_on_a_full_run():
    header = build_summary_rows([], FUNNELS)[0]
    assert all('не проверялся' not in str(cell) for cell in header)


def test_summary_still_counts_and_sums_when_a_class_is_marked():
    """Пометка — только слово в заголовке: числа и инвариант полноты целы."""
    findings = [finding(1), finding(17), finding(9, funnel='—')]
    rows = build_summary_rows(findings, FUNNELS, not_evaluated={9, 17})
    header = rows[0]
    cols = [i for i, name in enumerate(header) if str(name).startswith('Класс ')]
    assert sum(row[i] for row in rows[1:-1] for i in cols) == len(findings)


@pytest.mark.parametrize('cls', sorted(REGISTRY_ONLY_CLASSES))
def test_every_registry_only_class_is_marked_in_the_summary(cls):
    header = build_summary_rows([], FUNNELS, not_evaluated=REGISTRY_ONLY_CLASSES)[0]
    assert f'Класс {cls} (не проверялся)' in header
