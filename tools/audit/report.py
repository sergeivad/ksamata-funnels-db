#!/usr/bin/env python3
"""Запись карты расхождений в XLSX.

Лист на каждый класс заводится всегда, даже пустой: «ноль находок» должен
быть виден явно, иначе отсутствие листа читается как забытая проверка.

Но и пустой лист сам по себе двусмыслен: «проверили, чисто» и «не проверяли»
выглядят одинаково. Прогон без реестра предложений GetCourse (`--no-api` либо
пустой ответ API) оставляет шесть классов несчитанными, а ещё двум выключает
фильтр отставки. Отчёт обязан сказать об этом на том же листе, где читатель
видит число, — не в инструкции к отчёту: на инструкцию в этом пакете уже
полагались, и ноль совпадений месяц читался как «расхождений нет» (фаза 15).

Кто именно слепнет — знает findings.py (REGISTRY_ONLY_CLASSES и
REGISTRY_FILTERED_CLASSES), там же, где заведены сами классы.
"""

from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db_source import label_of
from findings import CLASS_TITLES

SUMMARY_SHEET = 'Сводка'
SOURCES_SHEET = 'Источники'

# 9 из 15 классов (2, 5, 7, 8, 9, 10, 11, 12, 14) — находки уровня
# предложения/оси/ключа: у них структурно нет одной воронки, item.funnel
# ставится в '—'. Туда же попадают сводные строки класса 1 по массовому
# отсутствующему тегу. Такие находки, а также находки с меткой воронки, которой
# нет среди funnels (опечатка/устаревшая метка), собираются в эту строку —
# иначе они бесследно пропадали бы из сводки.
UNASSIGNED_LABEL = '— (без воронки)'
TOTAL_LABEL = 'Всего'

CLASS_HEADERS = [
    'Воронка', 'Тип', 'Находка', 'Подробности',
    'Свидетельство', 'Первое наблюдение', 'Последнее наблюдение',
    'Заказов', 'Решение',
]

# Пометки листов. Текст полный и самодостаточный: читатель мог получить файл
# ссылкой и не видеть ни команды запуска, ни терминала.
NOT_EVALUATED_NOTE = (
    'ПРОВЕРКА НЕ ВЫПОЛНЯЛАСЬ: реестр предложений GetCourse не прочитан '
    '(прогон с --no-api либо пустой ответ API). Пустой список ниже НЕ значит '
    '«расхождений нет» — классу нечего было сравнивать. Режим прогона — на '
    'листе «Источники».'
)
DEGRADED_NOTE = (
    'ПРОГОН БЕЗ РЕЕСТРА GetCourse: выключен фильтр «в текущем реестре этого '
    'уже нет», поэтому в список могли попасть находки по разметке прошлого, '
    'которые полный прогон гасит. Число ниже — верхняя оценка, не итог. '
    'Режим прогона — на листе «Источники».'
)
UNEVALUATED_HEADER_SUFFIX = ' (не проверялся)'

HEADER_FILL = PatternFill('solid', fgColor='DDDDDD')
WARNING_FILL = PatternFill('solid', fgColor='FFF2CC')
WARNING_FONT = Font(bold=True, color='9C5700')
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)


def build_summary_rows(findings, funnels, not_evaluated=frozenset()):
    """Строка на воронку, колонка на класс. Первая строка — заголовок.

    `not_evaluated` — классы, которые в этом прогоне не считались. У них
    помечается ЗАГОЛОВОК колонки, а числа остаются нулями: сводка — таблица,
    и «н/д» вместо нуля сломал бы и сумму, и инвариант полноты ниже. Слово
    стоит вплотную к числу, потому что читатель смотрит именно туда.

    После строк воронок следуют две агрегирующие строки:

    - `UNASSIGNED_LABEL` — находки, чья `item.funnel` не совпадает ни с
      одной меткой из `funnels`. Это не только буквальный '—' (9 из 15
      классов ставят его структурно), но и любая метка воронки, которой
      среди `funnels` нет — иначе такая дыра осталась бы просто уже.
    - `TOTAL_LABEL` — сумма по каждому классу и общий итог.

    Инвариант: сумма всех чисел по классам во всех строках, кроме
    итоговой, равна общему числу находок — ни одна находка не теряется.
    """
    classes = sorted(CLASS_TITLES)
    header = (['Воронка', 'Продукт', 'Статус']
              + [f'Класс {c}' + (UNEVALUATED_HEADER_SUFFIX if c in not_evaluated else '')
                 for c in classes]
              + ['Всего'])

    counts = defaultdict(lambda: defaultdict(int))
    for item in findings:
        counts[item.funnel][item.cls] += 1

    known_labels = {label_of(row) for row in funnels}

    rows = [header]
    for row in funnels:
        label = label_of(row)
        per_class = [counts[label][c] for c in classes]
        rows.append([label, row.product_name, row.status] + per_class + [sum(per_class)])

    unassigned_per_class = [0] * len(classes)
    for label, by_cls in counts.items():
        if label in known_labels:
            continue
        for idx, cls in enumerate(classes):
            unassigned_per_class[idx] += by_cls[cls]
    rows.append([UNASSIGNED_LABEL, '', ''] + unassigned_per_class + [sum(unassigned_per_class)])

    totals_per_class = [0] * len(classes)
    for data_row in rows[1:]:
        for idx in range(len(classes)):
            totals_per_class[idx] += data_row[3 + idx]
    rows.append([TOTAL_LABEL, '', ''] + totals_per_class + [sum(totals_per_class)])

    return rows


def _write_sheet(ws, title, headers, rows, note=None):
    """Строка 1 — название листа, 2 — пометка (или пусто), 3 — заголовки,
    4+ — данные. Пометка занимает УЖЕ существовавшую пустую строку: номера
    строк — часть договора с читателем и с тестами, сдвигать их нельзя.
    """
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.append([])
    if note:
        ws['A2'] = note
        ws['A2'].font = WARNING_FONT
        ws['A2'].fill = WARNING_FILL
    ws.append(headers)
    for cell in ws[3]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append(row)
    ws.freeze_panes = 'A4'
    for index, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = 24


def write_report(path, findings, funnels, sources,
                 not_evaluated=frozenset(), degraded=frozenset()):
    """`not_evaluated` — классы, которые не считались вовсе (их пустота ничего
    не значит); `degraded` — считались, но с завышенным числом находок.
    Оба множества приходят из findings.unevaluated_classes/degraded_classes,
    то есть из данных прогона, а не из флага командной строки.
    """
    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = SUMMARY_SHEET
    rows = build_summary_rows(findings, funnels, not_evaluated)
    summary.append(rows[0])
    for cell in summary[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for row in rows[1:]:
        summary.append(row)
    summary.freeze_panes = 'B2'

    by_class = defaultdict(list)
    for item in findings:
        by_class[item.cls].append(item)

    for cls in sorted(CLASS_TITLES):
        ws = wb.create_sheet(f'Класс {cls}')
        body = [
            [
                item.funnel, item.tag_type, item.subject, item.detail,
                item.evidence, item.first_seen, item.last_seen, item.deals, None,
            ]
            for item in by_class.get(cls, [])
        ]
        note = (NOT_EVALUATED_NOTE if cls in not_evaluated
                else DEGRADED_NOTE if cls in degraded else None)
        _write_sheet(ws, f'Класс {cls}. {CLASS_TITLES[cls]}', CLASS_HEADERS, body, note)

    ws = wb.create_sheet(SOURCES_SHEET)
    body = [[s.get('kind', ''), s.get('name', ''), s.get('detail', '')] for s in sources]
    _write_sheet(ws, 'Источники прогона', ['Тип', 'Имя', 'Подробности'], body)

    wb.save(path)
