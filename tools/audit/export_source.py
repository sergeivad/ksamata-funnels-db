#!/usr/bin/env python3
"""Поиск и разбор выгрузок deal_export.

Ключевое свойство, проверенное на данных: «Теги предложений» вычисляются
в МОМЕНТ ВЫГРУЗКИ, а не хранятся на заказе. Поэтому единица наблюдения —
пара (заказ, дата файла), и дедупликация идёт по ней, а не по одному
id заказа: один заказ в двух выгрузках это два наблюдения, и именно их
расхождение является сигналом дрейфа.
"""

import csv
import datetime
import os
import re
import warnings
from dataclasses import dataclass

import openpyxl

from normalize import parse_tagset

TAGS_COLUMN = 'Теги предложений'
DEAL_ID_COLUMN = 'ID заказа'
CREATED_COLUMN = 'Дата создания'

FILE_PREFIX = 'deal_export'
DATE_RE = re.compile(r'(\d{4})-(\d{2})-(\d{2})')


@dataclass(frozen=True)
class Observation:
    deal_id: str
    tags: frozenset
    file_name: str
    file_date: datetime.date
    deal_created: str


def file_date_from_name(name):
    match = DATE_RE.search(name)
    if not match:
        return None
    try:
        return datetime.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _read_header(path):
    if path.lower().endswith('.csv'):
        with open(path, encoding='utf-8-sig', newline='') as fh:
            return next(csv.reader(fh, delimiter=';'), [])
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(c) for c in row if c is not None]
    finally:
        wb.close()


def has_tags_column(path):
    """Отсеивает срезы *_utm с 13 колонками — колонки тегов там нет.

    Возвращает False в двух разных по природе случаях, но предупреждает
    только об одном из них:
    - файл прочитан, колонки тегов в заголовке нет — штатный, ожидаемый
      исход (срез *_utm), молчим;
    - файл прочитать не удалось вовсе (битый архив, ошибка прав, неожиданный
      формат) — аномалия: файл тихо выпадает из охвата карты расхождений,
      поэтому предупреждаем с именем файла и причиной.
    """
    try:
        header = _read_header(path)
    except Exception as exc:
        warnings.warn(
            f'export_source: не удалось прочитать {path} — файл исключён из охвата ({exc!r})',
            stacklevel=2,
        )
        return False
    return TAGS_COLUMN in header


def _classify_candidates(directory, since):
    """Проходит по каталогу и классифицирует каждый файл-кандидат deal_export.

    Кандидат — файл с именем на FILE_PREFIX, не служебная блокировка ('~$'),
    расширение .csv/.xlsx. Отдаёт (полный_путь, category), где category:
    'ok' — попадает в охват карты расхождений;
    'too_old' — дата файла раньше since (или её вообще не удалось разобрать);
    'no_tags' — колонки «Теги предложений» в файле нет (например, срез *_utm).

    Общая точка правды для discover_export_files (только 'ok') и
    discover_export_files_with_stats (все три категории, с подсчётом).
    """
    for name in sorted(os.listdir(directory)):
        if name.startswith('~$') or not name.startswith(FILE_PREFIX):
            continue
        if not name.lower().endswith(('.csv', '.xlsx')):
            continue
        full = os.path.join(directory, name)
        file_date = file_date_from_name(name)
        if file_date is None or file_date < since:
            yield full, 'too_old'
        elif has_tags_column(full):
            yield full, 'ok'
        else:
            yield full, 'no_tags'


def discover_export_files(directory, since):
    return [full for full, category in _classify_candidates(directory, since)
            if category == 'ok']


def discover_export_files_with_stats(directory, since):
    """Как discover_export_files, но вдобавок считает, что и почему отброшено.

    Возвращает (files, stats), где files — тот же список отобранных путей,
    что и у discover_export_files (для того же входа), а stats — словарь:
      - 'total_candidates' — всего файлов-кандидатов deal_export,
      - 'selected' — попало в охват (== len(files)),
      - 'excluded_too_old' — отброшено как более старые, чем since (включая
        файлы без разбираемой даты в имени),
      - 'excluded_no_tags_column' — отброшено как не содержащие колонку
        «Теги предложений».
    Нужно для листа «Источники»: без этого читатель отчёта не может оценить
    полноту выборки.
    """
    files = []
    stats = {
        'total_candidates': 0,
        'selected': 0,
        'excluded_too_old': 0,
        'excluded_no_tags_column': 0,
    }
    for full, category in _classify_candidates(directory, since):
        stats['total_candidates'] += 1
        if category == 'ok':
            files.append(full)
            stats['selected'] += 1
        elif category == 'too_old':
            stats['excluded_too_old'] += 1
        else:
            stats['excluded_no_tags_column'] += 1
    return files, stats


def _rows(path):
    """Отдаёт словари {колонка: значение} независимо от формата файла."""
    if path.lower().endswith('.csv'):
        with open(path, encoding='utf-8-sig', newline='') as fh:
            for row in csv.DictReader(fh, delimiter=';'):
                yield row
        return

    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb.active
        stream = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else '' for c in next(stream, ())]
        for values in stream:
            yield {header[i]: values[i] for i in range(min(len(header), len(values)))}
    finally:
        wb.close()


def read_observations(path):
    name = os.path.basename(path)
    file_date = file_date_from_name(name)
    if file_date is None:
        return []

    result = []
    for row in _rows(path):
        deal_id = row.get(DEAL_ID_COLUMN)
        if deal_id is None or str(deal_id).strip() == '':
            continue
        tags = parse_tagset(row.get(TAGS_COLUMN))
        if not tags:
            continue
        result.append(
            Observation(
                deal_id=str(deal_id).strip(),
                tags=tags,
                file_name=name,
                file_date=file_date,
                deal_created=str(row.get(CREATED_COLUMN) or '').strip(),
            )
        )
    return result


def load_observations(files):
    """Дедуп по (заказ, дата файла) — НЕ по одному заказу.

    Две выгрузки одной даты с одним и тем же заказом дают одно наблюдение;
    выгрузки разных дат — два, даже если наборы тегов совпали.
    """
    seen = {}
    for path in files:
        for obs in read_observations(path):
            seen.setdefault((obs.deal_id, obs.file_date), obs)
    return sorted(seen.values(), key=lambda o: (o.file_date, o.deal_id))
