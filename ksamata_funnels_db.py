#!/usr/bin/env python3
"""
Создание и заполнение SQLite базы данных автоворонок Ксамата.
Источники: 3 Excel-файла + ручные overrides из patch_v6.py
"""

import sqlite3
import openpyxl
import re
import os
import json
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'ksamata_funnels.db')
FILE1 = os.path.join(BASE_DIR, 'Ссылки для сбора статы.xlsx')
FILE2 = os.path.join(BASE_DIR, 'Воронки ссылки-2.xlsx')
FILE3 = os.path.join(BASE_DIR, 'source_data', 'salebot_calculator_results_20260329-143800.xlsx')

# ============================================================
# 1. SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS sources (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS tags (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS products (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS contractors (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS funnels (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    num             INTEGER NOT NULL UNIQUE,
    source_id       INTEGER NOT NULL REFERENCES sources(id),
    product_id      INTEGER NOT NULL REFERENCES products(id),
    contractor_id   INTEGER NOT NULL REFERENCES contractors(id),
    variant         TEXT NOT NULL DEFAULT '',
    product_name    TEXT NOT NULL DEFAULT '',
    landing_url     TEXT DEFAULT '',
    start_date      TEXT DEFAULT '',
    block_name      TEXT DEFAULT '',
    sheet_name      TEXT DEFAULT '',
    tag_19_raw      TEXT DEFAULT '',
    tag_15_raw      TEXT DEFAULT '',
    reg_tags_raw    TEXT DEFAULT '',
    dash_sales_url  TEXT DEFAULT '',
    dash_pereliv_url TEXT DEFAULT '',
    regi_total_url  TEXT DEFAULT '',
    regi_15_url     TEXT DEFAULT '',
    regi_19_url     TEXT DEFAULT '',
    regi_notime_url TEXT DEFAULT '',
    predspisok_url  TEXT DEFAULT '',
    room_ids_json   TEXT DEFAULT '{}',
    bothelp_condition TEXT DEFAULT '',
    created_at      TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at      TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS funnel_tags (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    funnel_id   INTEGER NOT NULL REFERENCES funnels(id) ON DELETE CASCADE,
    tag_id      INTEGER NOT NULL REFERENCES tags(id),
    tag_type    TEXT NOT NULL CHECK(tag_type IN ('reg', 'time_19', 'time_15')),
    position    INTEGER NOT NULL DEFAULT 0,
    UNIQUE(funnel_id, tag_id, tag_type)
);

CREATE TABLE IF NOT EXISTS funnel_days (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    funnel_id    INTEGER NOT NULL REFERENCES funnels(id) ON DELETE CASCADE,
    time_slot    TEXT NOT NULL CHECK(time_slot IN ('19', '15')),
    day_num      INTEGER NOT NULL CHECK(day_num BETWEEN 1 AND 5),
    room_id_f1   TEXT DEFAULT '',
    gc_room      TEXT DEFAULT '',
    web_room     TEXT DEFAULT '',
    replay_url   TEXT DEFAULT '',
    web_replay   TEXT DEFAULT '',
    sales_page   TEXT DEFAULT '',
    sales_note   TEXT DEFAULT '',
    tariffs      TEXT DEFAULT '',
    oto          TEXT DEFAULT '',
    bonuses      TEXT DEFAULT '',
    mission      TEXT DEFAULT '',
    mission_type TEXT DEFAULT '',
    meditation   TEXT DEFAULT '',
    dojim_note   TEXT DEFAULT '',
    UNIQUE(funnel_id, time_slot, day_num)
);

CREATE TABLE IF NOT EXISTS salebot_configs (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    funnel_id   INTEGER NOT NULL REFERENCES funnels(id) ON DELETE CASCADE,
    time_slot   TEXT NOT NULL CHECK(time_slot IN ('19', '15')),
    condition   TEXT DEFAULT '',
    calculator  TEXT DEFAULT '',
    UNIQUE(funnel_id, time_slot)
);

CREATE INDEX IF NOT EXISTS idx_funnel_tags_funnel ON funnel_tags(funnel_id);
CREATE INDEX IF NOT EXISTS idx_funnel_tags_tag    ON funnel_tags(tag_id);
CREATE INDEX IF NOT EXISTS idx_funnel_days_funnel ON funnel_days(funnel_id);
CREATE INDEX IF NOT EXISTS idx_salebot_funnel     ON salebot_configs(funnel_id);
CREATE INDEX IF NOT EXISTS idx_funnels_product    ON funnels(product_id);
CREATE INDEX IF NOT EXISTS idx_funnels_contractor ON funnels(contractor_id);

CREATE TRIGGER IF NOT EXISTS trg_funnels_updated
AFTER UPDATE ON funnels
BEGIN
    UPDATE funnels SET updated_at = datetime('now') WHERE id = NEW.id;
END;
"""

# ============================================================
# 2. OVERRIDES (from manual edits synced in patch_v6)
# ============================================================

landing_overrides = {
    1: 'https://lp.ksamata.ru/dtx-yo',
    2: 'http://lp.ksamata.ru/izh-yo',
    3: 'https://lp.ksamata.ru/rd-yo',
    4: 'http://lp.ksamata.ru/izh-yo',
}

tag_overrides = {
    1: {'tag_19': '', 'tag_15': ''},
    11: {'tag_19': 'ДБО, ВК NR, ВК NR ВК, ВК NR IS, t=19', 'tag_15': 'ДБО, ВК NR, ВК NR ВК, ВК NR IS, t=15'},
    15: {'tag_19': 'ДБО, ВК NR, ВК NR ВК, ВК NR МП, t=19', 'tag_15': 'ДБО, ВК NR, ВК NR ВК, ВК NR МП, t=15'},
    25: {'tag_19': 'СВС, сосуды, Яндекс Ретаргет, t=19', 'tag_15': 'СВС, сосуды, Яндекс Ретаргет, t=15'},
    26: {'tag_19': 'ДБО, Яндекс Ретаргет, t=19', 'tag_15': 'ДБО, Яндекс Ретаргет, t=15'},
}

block_overrides = {
    15: {'regi_total_override': 'https://gc.ksamata.ru/pl/user/user/index?uc%5Bsegment_id%5D=0&uc%5Brule_string%5D=%7B%22type%22%3A%22user_hasdealrule%22%2C%22inverted%22%3A0%2C%22params%22%3A%7B%22linkedRule%22%3A%7B%22type%22%3A%22andrule%22%2C%22inverted%22%3A0%2C%22params%22%3A%7B%22mode%22%3A%22and%22%2C%22children%22%3A%5B%7B%22type%22%3A%22deal_created_at%22%2C%22inverted%22%3A0%2C%22params%22%3A%7B%22value%22%3A%7B%22from%22%3A%2215.07.2024%22%2C%22to%22%3A%2215.07.2024%22%2C%22toNDays%22%3Anull%2C%22fromNDays%22%3Anull%2C%22dateType%22%3A%22prev_day%22%2C%22withTime%22%3Afalse%7D%2C%22valueMode%22%3Anull%7D%7D%2C%7B%22type%22%3A%22deal_offer_id%22%2C%22inverted%22%3A0%2C%22params%22%3A%7B%22value%22%3A%7B%22selected_id%22%3A%5B8033465%5D%2C%22selected_tags%22%3A%5B%5D%2C%22all_object_with_tags%22%3Afalse%7D%2C%22valueMode%22%3Anull%7D%7D%5D%7D%7D%2C%22countCondition%22%3A%7B%22checker%22%3A%22nlt%22%2C%22numval%22%3A%22%22%7D%7D%2C%22maxSize%22%3A%22%22%7D&formParams%5Bclarity_uid%5D=wIMVGIqMM-l54Ric6Z1s6gW9Zhwq0Vn1'},
    26: {'block_name': '[ДБО Яндекс Ретаргет] НОВАЯ ЦЕНА', 'sheet': 'ДБО'},
}

reg_tags = {
    1: 'Регистрация, Ютуб органика',
    2: 'Регистрация, Жизнь(СВС)',
    3: 'Регистрация, РД',
    4: 'Регистрация, Ютуб органика, Ютуб О, ГП',
    5: 'Регистрация, Ютуб Реклама, НИМБ, РД',
    6: 'Регистрация, яндекс, Детокс, РСЯ',
    7: 'Регистрация, РСЯ, РД',
    8: 'Регистрация, Яндекс Реклама новый ленд, ЖКТ, РСЯ',
    9: 'Регистрация, Яндекс Реклама новый ленд, ЩЖ, РСЯ',
    10: 'Регистрация, сосуды, РСЯ',
    11: 'Регистрация, ВК Сода, РД, ВК NR ВК, ВК NR IS, ВК NR',
    12: 'Регистрация, ЖКТ, ВК NR ВК, ВК NR',
    13: 'Регистрация, ЖКТ, ВК NR IS, ВК NR',
    14: 'Регистрация, ЖКТ, ВК NR МП, ВК NR',
    15: 'Регистрация, РД, ВК NR МП, ВК NR ВК, ВК NR',
    16: 'Регистрация, Детокс, ВК БАИНГ, ВК',
    17: 'Регистрация, МАКС, РД, FAQ',
    18: 'Регистрация, РД, ВК HT, ВК',
    19: 'Регистрация, Детокс, ВК HT',
    20: 'Регистрация, Детокс, ВК БАИНГ, ВК',
    21: 'Регистрация, Яндекс Реклама, Жизнь(СВС)',
    22: 'Регистрация, Яндекс Реклама, яндекс, Детокс',
    23: 'Регистрация, Яндекс Реклама, РД',
    24: 'Регистрация, Яндекс Реклама, яндекс, Детокс',
    25: 'Регистрация, Яндекс Ретаргет, Жизнь(СВС)',
    26: 'Регистрация, Яндекс Ретаргет, РД, яндекс',
}

start_dates = {
    1: '01.08.2024', 2: '01.08.2024', 3: '26.11.2024', 4: '27.11.2024',
    5: '11.02.2026', 6: '29.05.2024', 7: '16.10.2024', 8: '01.01.2025',
    9: '11.08.2025', 10: '27.08.2025', 11: '18.09.2025', 12: '08.09.2025',
    13: '04.10.2025', 14: '11.11.2025', 15: '03.02.2026', 16: '03.03.2026',
    17: '08.03.2026', 18: '01.05.2025', 19: '01.12.2025', 20: '12.02.2026',
    21: '01.12.2021', 22: '01.12.2021', 23: '01.12.2021', 24: '01.12.2021',
    25: '01.12.2021', 26: '01.12.2021',
}

# Day-level overrides for funnels with manually corrected room URLs in v5
day_overrides = {
    25: {
        ("19", 1): {'gc_room': 'https://gc.ksamata.ru/cvc1-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc1-19-yanr', 'sales_page': 'https://t.ksamata.ru/cvc/tarif-19-yanr', 'tariffs': 'https://gc.ksamata.ru/cvc/tarif/curator-19-yanr', 'oto': 'https://gc.ksamata.ru/cvc/oto-19-yanr', 'bonuses': 'https://gc.ksamata.ru/av/cvc/bonus1', 'mission_gc': 'https://gc.ksamata.ru/pl/tasks/mission/update?id=2196056'},
        ("19", 2): {'gc_room': 'https://gc.ksamata.ru/cvc2-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc2-19-yanr', 'sales_page': 'https://t.ksamata.ru/cvc/tarifz-19-yanr', 'sales_note': 'тарифы с записью ГЛАВНОГО занятия', 'bonuses': 'https://gc.ksamata.ru/av/cvc/bonus2', 'mission_gc': 'https://gc.ksamata.ru/pl/tasks/mission/process?id=2177131'},
        ("19", 3): {'gc_room': 'https://gc.ksamata.ru/cvc3-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc3-19-yanr', 'bonuses': 'https://gc.ksamata.ru/av/cvc/program', 'mission_gc': 'https://salebot.pro/projects/98250/messages?sheet_id=338826'},
        ("19", 4): {'gc_room': 'https://gc.ksamata.ru/cvc4-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc4-19-yanr'},
        ("19", 5): {'gc_room': 'https://gc.ksamata.ru/cvc5-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc5-19-yanr'},
        ("15", 1): {'gc_room': 'https://gc.ksamata.ru/cvc1-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc1-15-yanr', 'sales_page': 'https://t.ksamata.ru/cvc/tarif-15-yanr', 'tariffs': 'https://gc.ksamata.ru/cvc/tarif/curator-15-yanr', 'oto': 'https://gc.ksamata.ru/cvc/oto-15-yanr'},
        ("15", 2): {'gc_room': 'https://gc.ksamata.ru/cvc2-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc2-15-yanr', 'sales_page': 'https://t.ksamata.ru/cvc/tarifz-15-yanr', 'sales_note': 'тарифы с записью ГЛАВНОГО занятия'},
        ("15", 3): {'gc_room': 'https://gc.ksamata.ru/cvc3-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc3-15-yanr'},
        ("15", 4): {'gc_room': 'https://gc.ksamata.ru/cvc4-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc4-15-yanr'},
        ("15", 5): {'gc_room': 'https://gc.ksamata.ru/cvc5-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/cvc5-15-yanr'},
    },
    26: {
        ("19", 1): {'gc_room': 'https://gc.ksamata.ru/dbo1-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo1-19-yanr', 'sales_page': 'https://t.ksamata.ru/spb/tarif-1yr', 'tariffs': 'https://t.ksamata.ru/dbo/tarif-19-yanr', 'oto': 'https://gc.ksamata.ru/dbo/oto-19-yanr', 'mission_gc': 'https://gc.ksamata.ru/pl/tasks/mission/process?id=2107799'},
        ("19", 2): {'gc_room': 'https://gc.ksamata.ru/dbo2-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo2-19-yanr', 'sales_page': 'https://t.ksamata.ru/spb/tarif-2yr', 'tariffs': 'https://t.ksamata.ru/dbo/tarifz-19-yanr', 'mission_gc': 'https://gc.ksamata.ru/pl/tasks/mission/update?id=2087552'},
        ("19", 3): {'gc_room': 'https://gc.ksamata.ru/dbo3-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo3-19-yanr', 'sales_page': 'https://t.ksamata.ru/spb/tarif-zyr', 'sales_note': 'тарифы с записью', 'mission_gc': 'https://salebot.pro/projects/98250/messages?sheet_id=303052'},
        ("19", 4): {'gc_room': 'https://gc.ksamata.ru/dbo4-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo4-19-yanr'},
        ("19", 5): {'gc_room': 'https://gc.ksamata.ru/dbo5-19-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo5-19-yanr'},
        ("15", 1): {'gc_room': 'https://gc.ksamata.ru/dbo1-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo1-15-yanr', 'sales_page': 'https://t.ksamata.ru/spb/tarif-yr1', 'tariffs': 'https://t.ksamata.ru/dbo/tarif-15-yanr', 'oto': 'https://gc.ksamata.ru/dbo/oto-15-yanr'},
        ("15", 2): {'gc_room': 'https://gc.ksamata.ru/dbo2-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo2-15-yanr', 'sales_page': 'https://t.ksamata.ru/spb/tarif-yr2', 'tariffs': 'https://t.ksamata.ru/dbo/tarifz-15-yanr'},
        ("15", 3): {'gc_room': 'https://gc.ksamata.ru/dbo3-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo3-15-yanr', 'sales_page': 'https://t.ksamata.ru/spb/tarif-yrz', 'sales_note': 'тарифы с записью'},
        ("15", 4): {'gc_room': 'https://gc.ksamata.ru/dbo4-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo4-15-yanr'},
        ("15", 5): {'gc_room': 'https://gc.ksamata.ru/dbo5-15-yanr', 'web_room': 'https://web.ksamatacenter.com/room/dbo5-15-yanr'},
    },
}

# Salebot overrides for funnels with manually corrected conditions
salebot_overrides = {
    19: {
        '19': {'condition': 'boo-vkns19', 'calculator': 'choose_time = 19\nboo_ht_date = #{current_date}'},
        '15': {'condition': 'boo-vkns15', 'calculator': 'choose_time = 15\nboo_ht_date = #{current_date}'},
    },
    25: {
        '19': {'condition': 'cvc-yanr-19', 'calculator': 'ct_av = 19\ncontr_id = "yanr"\n\ncvc_yanr_date = #{current_date}'},
        '15': {'condition': 'cvc-yanr-15', 'calculator': 'ct_av = 15\ncontr_id = "yanr"\n\ncvc_yanr_date = #{current_date}'},
    },
    26: {
        '19': {'condition': '', 'calculator': 'dbo-date = #{current_date}\n\nct_av = 19\n\ncontr_id = "yanr"'},
        '15': {'condition': '', 'calculator': 'dbo-date = #{current_date}\n\nct_av = 15\n\ncontr_id = "yanr"'},
    },
}

# BotHelp conditions: funnel_num -> condition string
bothelp_conditions = {
    5: 'dbo_ytn_date',
    6: 'boo_rsya_date',
    7: 'dbo_yr_date',
    8: 'zkt_rsya_date',
    9: 'shzh_rsya_date',
    10: 'cvc_vknimb_date',
    11: 'dbo_nr_date',
    12: 'zkt_nr_date',
    13: 'zkt_isnr_date',
    14: 'zkt_mpnr_date',
    15: 'dbo_nrmp_date',
    17: 'dbo_omax_date',
    18: 'dbo_ht_date',
    19: 'boo_ht_date',
    21: 'cvc_yan_date',
    22: 'boo_y_date',
    23: 'dbo_y_date',
    26: 'dbo_yanr_date',
}

# Funnel mapping: num -> (product, contractor, variant)
funnel_mapping = {
    1:  ('БОО', 'Органика', 'ЮТУБ'),
    2:  ('СВС', 'Органика', 'ЮТУБ'),
    3:  ('ДБО', 'Органика', 'ЮТУБ'),
    4:  ('ГП', 'Органика', 'ЮТУБ'),
    5:  ('ДБО', 'НИМБ', 'ЮТУБ'),
    6:  ('БОО', 'НИМБ', 'РСЯ'),
    7:  ('ДБО', 'НИМБ', 'РСЯ'),
    8:  ('ЖКТ', 'НИМБ', 'РСЯ'),
    9:  ('ЩЖ', 'НИМБ', 'РСЯ'),
    10: ('СВС', 'НИМБ', 'РСЯ'),
    11: ('ДБО', 'NR', 'ВК'),
    12: ('ЖКТ', 'NR', 'ВК'),
    13: ('ЖКТ', 'NR', 'IS'),
    14: ('ЖКТ', 'NR', 'МП'),
    15: ('ДБО', 'NR', 'МП'),
    16: ('БОО', 'NR', 'IS'),
    17: ('ДБО', 'FAQ', 'MAX'),
    18: ('ДБО', 'HT', 'ВК'),
    19: ('БОО', 'HT', 'ВК'),
    20: ('БОО', 'ВК БАИНГ', ''),
    21: ('СВС', 'Алексей', 'Яндекс Реклама'),
    22: ('БОО', 'Алексей', 'Яндекс Реклама'),
    23: ('ДБО', 'Алексей', 'Яндекс Реклама'),
    24: ('БОО', 'Алексей', 'Яндекс Ретаргет'),
    25: ('СВС', 'Алексей', 'Яндекс Ретаргет'),
    26: ('ДБО', 'Алексей', 'Яндекс Ретаргет'),
}

# Tag raw string override (extra space fix)
tag_raw_overrides = {
    21: {'tag_15': 'сосуды, СВС , Яндекс Реклама, t=15'},
}

# ============================================================
# 3. HELPERS (same logic as patch_v6.py)
# ============================================================

def sv(v):
    return str(v).strip() if v and str(v).strip() != 'None' else ''

def prepend_autofunnel(raw):
    """Prepend 'автоворонки' to a time-tag raw string if not already present.
    Empty strings are left empty (funnels without time tags shouldn't get one)."""
    if not raw:
        return raw
    parts = [t.strip() for t in raw.split(',')]
    if 'автоворонки' in parts:
        return raw
    return 'автоворонки, ' + raw

def cond_to_key(cond):
    if not cond: return '', ''
    c = str(cond).strip()
    time = ''
    if c.endswith('19') or c.endswith('19-old'):
        time = '19'
    elif c.endswith('15') or c.endswith('15-old'):
        time = '15'
    key = re.sub(r'(?:19|15)(?:-old)?$', '', c).rstrip('-')
    return key, time

def room_to_key(room_id):
    if not room_id: return ''
    r = re.sub(r'^\d+', '', room_id)
    r = re.sub(r'^([a-z]+)\d+-', r'\1-', r)
    r = re.sub(r'-(?:15|19)-', '-', r)
    r = r.replace('-old', '')
    return r

def norm_tag(t, time):
    if not t: return t
    t = re.sub(r'[,\s]*(19:00|15:00|t=19|t=15)\s*$', '', t).rstrip(', ').rstrip()
    return t + f', t={time}'

def is_tag(val):
    if not val: return False
    s = str(val).strip()
    if not s or s in ('None', '15:00:00', '19:00:00'): return False
    if s.startswith('Отключена') or s.startswith('МОЖНО') or s.startswith('СВОБОДНАЯ'): return False
    return ',' in s or s.lower().startswith('тег')

def parse_tag_string(raw):
    """Split comma-separated tag string, filter out time markers."""
    if not raw:
        return []
    parts = [t.strip() for t in raw.split(',')]
    return [p for p in parts if p and not re.match(r'^t=\d+$', p)]

def format_sb_calculator(calc):
    """Filter calculator text: remove site/web/concat variables."""
    if not calc:
        return ''
    lines = calc.split('\n')
    filtered = []
    for l in lines:
        s = l.strip()
        if not s:
            continue
        if s.startswith('site ') or s.startswith('site='):
            continue
        if s.startswith('web ') or s.startswith('web='):
            continue
        if 'concat(' in s and ('gc.ksamata.ru' in s or 't.ksamata.ru' in s):
            continue
        filtered.append(l)
    return '\n'.join(filtered).strip()

# ============================================================
# 4. EXTRACT DATA FROM EXCEL FILES
# ============================================================

def load_salebot(wb3):
    ws3 = wb3['results']
    salebot_data = {}
    for r in range(2, ws3.max_row + 1):
        condition = str(ws3.cell(r, 5).value or '').strip()
        calculator = str(ws3.cell(r, 6).value or '').strip()
        if calculator == 'None':
            calculator = ''
        status = str(ws3.cell(r, 8).value or '')
        if status != 'success':
            continue
        key, time = cond_to_key(condition)
        if not key:
            continue
        if key not in salebot_data:
            salebot_data[key] = {}
        if time:
            salebot_data[key][time] = {'condition': condition, 'calculator': calculator}
        else:
            salebot_data[key]['all'] = {'condition': condition, 'calculator': calculator}
    return salebot_data


def load_funnels(wb1, salebot_data):
    ws1 = wb1['Рабочие']
    funnels = []
    current_source = ""
    for row_num in range(5, 31):
        row = [ws1.cell(row_num, c).value for c in range(1, 22)]
        num = row[0]
        if not num:
            continue
        source = row[1] if row[1] else current_source
        if row[1]:
            current_source = row[1]
        rooms = {}
        for day in range(5):
            rooms[f'd{day+1}_15'] = str(row[10 + day * 2]).strip() if row[10 + day * 2] else ""
            rooms[f'd{day+1}_19'] = str(row[11 + day * 2]).strip() if row[11 + day * 2] else ""
        funnels.append({
            'num': int(num), 'source': str(source).strip(),
            'name': str(row[2]).strip() if row[2] else "",
            'landing': str(row[3]).strip() if row[3] else "",
            'dash_sales': str(row[4]).strip() if row[4] else "",
            'dash_pereliv': str(row[5]).strip() if row[5] else "",
            'regi_total': str(row[6]).strip() if row[6] else "",
            'regi_15': str(row[7]).strip() if row[7] else "",
            'regi_19': str(row[8]).strip() if row[8] else "",
            'regi_notime': str(row[9]).strip() if row[9] else "",
            'rooms': rooms,
        })

    # Match salebot
    for f in funnels:
        room_19 = f['rooms'].get('d1_19', '')
        key = room_to_key(room_19)
        sb = salebot_data.get(key, {})
        f['sb_19'] = sb.get('19', sb.get('all', {}))
        f['sb_15'] = sb.get('15', {})

    return funnels


def find_block(wb2, room_id):
    if not room_id:
        return None, None
    for sn in wb2.sheetnames:
        ws = wb2[sn]
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 3).value
            if v and room_id in str(v):
                return ws, r
    return None, None


def extract_block(ws, anchor, sn):
    result = {
        'sheet': sn, 'block_name': '', 'tag_19': '', 'tag_15': '',
        'days_19': {}, 'days_15': {},
    }
    for off in range(-5, 0):
        r = anchor + off
        if r < 1:
            continue
        vb = ws.cell(r, 2).value
        if vb and '[' in str(vb):
            result['block_name'] = sv(vb)
            break

    rows_data = []
    for r in range(anchor, min(anchor + 50, ws.max_row + 1)):
        va = sv(ws.cell(r, 1).value)
        vb = sv(ws.cell(r, 2).value)
        vc = sv(ws.cell(r, 3).value)
        if vb and '[' in vb and r > anchor + 2:
            break
        if va in ('Отключена', 'МОЖНО УДАЛЯТЬ КОМНАТЫ', 'СВОБОДНАЯ') and r > anchor + 2:
            break
        rows_data.append({
            'row': r, 'a': va, 'b': vb, 'c': vc,
            'has_day': 'день' in vb if vb else False,
            'is_gc': bool(vc and 'gc.ksamata.ru' in vc),
            'is_web': bool(vc and 'web.ksamatacenter.com' in vc),
            'is_15': va == '15:00:00',
            'is_19': va == '19:00:00',
        })

    split_idx = None
    for i, rd in enumerate(rows_data):
        if rd['is_15'] and i > 2:
            split_idx = i
            break

    section_19 = rows_data[:split_idx] if split_idx else rows_data
    section_15 = rows_data[split_idx + 1:] if split_idx else []

    def parse_section(section):
        gc_rows = []
        web_rows = []
        tag = ''
        for rd in section:
            if not rd['has_day'] and not rd['is_gc'] and not rd['is_web']:
                if is_tag(rd['a']):
                    tag = rd['a']
                continue
            if rd['has_day']:
                if is_tag(rd['a']):
                    tag = rd['a']
                if rd['is_gc']:
                    gc_rows.append(rd['row'])
                elif rd['is_web']:
                    web_rows.append(rd['row'])
                elif rd['is_19'] or rd['is_15']:
                    gc_rows.append(rd['row'])
                elif not rd['c']:
                    web_rows.append(rd['row'])
                else:
                    gc_rows.append(rd['row'])
            elif rd['is_web']:
                web_rows.append(rd['row'])
            elif rd['is_gc']:
                gc_rows.append(rd['row'])
        return gc_rows, web_rows, tag

    gc_19_r, web_19_r, tag_19 = parse_section(section_19)
    gc_15_r, web_15_r, tag_15 = parse_section(section_15)

    result['tag_19'] = norm_tag(tag_19, '19')
    result['tag_15'] = norm_tag(tag_15, '15')

    def extract_gc(r):
        return {
            'gc_room': sv(ws.cell(r, 3).value),
            'replay': sv(ws.cell(r, 5).value),
            'sales_page': sv(ws.cell(r, 6).value),
            'sales_note': sv(ws.cell(r, 7).value),
            'tariffs': sv(ws.cell(r, 8).value),
            'oto': sv(ws.cell(r, 9).value),
            'bonuses': sv(ws.cell(r, 10).value),
            'mission_gc': sv(ws.cell(r, 11).value),
        }

    def extract_web(r):
        return {
            'web_room': sv(ws.cell(r, 3).value),
            'web_replay': sv(ws.cell(r, 5).value),
            'meditation': sv(ws.cell(r, 6).value),
            'dojim_note': sv(ws.cell(r, 7).value),
            'mission_web': sv(ws.cell(r, 11).value),
            'mission_type': sv(ws.cell(r, 13).value),
        }

    for i, r in enumerate(gc_19_r):
        result['days_19'][i + 1] = extract_gc(r)
    for i, r in enumerate(web_19_r):
        d = i + 1
        wd = extract_web(r)
        if d in result['days_19']:
            result['days_19'][d].update(wd)
        else:
            result['days_19'][d] = wd

    for i, r in enumerate(gc_15_r):
        result['days_15'][i + 1] = extract_gc(r)
    for i, r in enumerate(web_15_r):
        d = i + 1
        wd = extract_web(r)
        if d in result['days_15']:
            result['days_15'][d].update(wd)
        else:
            result['days_15'][d] = wd

    return result


# ============================================================
# 5. POPULATE DATABASE
# ============================================================

def populate(db_path):
    # Remove old DB if exists
    if os.path.exists(db_path):
        os.remove(db_path)

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(SCHEMA)

    # Load Excel data
    print("Loading Excel files...")
    wb1 = openpyxl.load_workbook(FILE1)
    wb2 = openpyxl.load_workbook(FILE2)
    wb3 = openpyxl.load_workbook(FILE3)

    salebot_data = load_salebot(wb3)
    print(f"  Salebot: {len(salebot_data)} keys")

    funnels = load_funnels(wb1, salebot_data)
    print(f"  Funnels: {len(funnels)}")

    # Extract blocks from file 2
    for f in funnels:
        room_19 = f['rooms'].get('d1_19', '')
        room_15 = f['rooms'].get('d1_15', '')
        ws, anchor = find_block(wb2, room_19)
        if not ws and room_15:
            ws, anchor = find_block(wb2, room_15)
        f['block'] = extract_block(ws, anchor, ws.title) if ws else {}

    # --- Insert sources ---
    source_names = sorted(set(f['source'] for f in funnels))
    source_ids = {}
    for name in source_names:
        cur = conn.execute("INSERT INTO sources(name) VALUES(?)", (name,))
        source_ids[name] = cur.lastrowid
    print(f"  Sources: {len(source_ids)}")

    # --- Insert products ---
    product_names = sorted(set(m[0] for m in funnel_mapping.values()))
    product_ids = {}
    for name in product_names:
        cur = conn.execute("INSERT INTO products(name) VALUES(?)", (name,))
        product_ids[name] = cur.lastrowid
    print(f"  Products: {len(product_ids)}")

    # --- Insert contractors ---
    contractor_names = sorted(set(m[1] for m in funnel_mapping.values()))
    contractor_ids = {}
    for name in contractor_names:
        cur = conn.execute("INSERT INTO contractors(name) VALUES(?)", (name,))
        contractor_ids[name] = cur.lastrowid
    print(f"  Contractors: {len(contractor_ids)}")

    # --- Collect all unique tags ---
    all_tag_names = set()
    for f in funnels:
        fnum = f['num']
        b = f.get('block', {})

        # Time tags (apply tag_raw_overrides too)
        tag_19_raw = tag_overrides.get(fnum, {}).get('tag_19', b.get('tag_19', ''))
        tag_15_raw = tag_overrides.get(fnum, {}).get('tag_15', b.get('tag_15', ''))
        tro = tag_raw_overrides.get(fnum, {})
        if 'tag_19' in tro:
            tag_19_raw = tro['tag_19']
        if 'tag_15' in tro:
            tag_15_raw = tro['tag_15']
        tag_19_raw = prepend_autofunnel(tag_19_raw)
        tag_15_raw = prepend_autofunnel(tag_15_raw)
        reg_raw = reg_tags.get(fnum, '')

        for raw in [tag_19_raw, tag_15_raw, reg_raw]:
            all_tag_names.update(parse_tag_string(raw))

    tag_ids = {}
    for name in sorted(all_tag_names):
        cur = conn.execute("INSERT INTO tags(name) VALUES(?)", (name,))
        tag_ids[name] = cur.lastrowid
    print(f"  Tags: {len(tag_ids)}")

    # --- Insert funnels + related data ---
    days_count = 0
    salebot_count = 0
    ftag_count = 0

    for f in funnels:
        fnum = f['num']
        b = f.get('block', {})
        bo = block_overrides.get(fnum, {})

        # Apply overrides
        landing = landing_overrides.get(fnum, f['landing'])
        block_name = bo.get('block_name', b.get('block_name', ''))
        sheet_name = bo.get('sheet', b.get('sheet', ''))
        regi_total = bo.get('regi_total_override', f['regi_total'])

        tag_19_raw = tag_overrides.get(fnum, {}).get('tag_19', b.get('tag_19', ''))
        tag_15_raw = tag_overrides.get(fnum, {}).get('tag_15', b.get('tag_15', ''))
        # Apply tag_raw_overrides (preserving exact text from v5, e.g. spacing)
        tro = tag_raw_overrides.get(fnum, {})
        if 'tag_19' in tro:
            tag_19_raw = tro['tag_19']
        if 'tag_15' in tro:
            tag_15_raw = tro['tag_15']
        tag_19_raw = prepend_autofunnel(tag_19_raw)
        tag_15_raw = prepend_autofunnel(tag_15_raw)
        reg_raw = reg_tags.get(fnum, '')

        # Convert date from DD.MM.YYYY to YYYY-MM-DD for SQLite
        sd = start_dates.get(fnum, '')
        if sd:
            try:
                sd = datetime.strptime(sd, '%d.%m.%Y').strftime('%Y-%m-%d')
            except ValueError:
                pass

        # Room IDs from File 1 as JSON
        room_ids = {}
        for day in range(1, 6):
            for slot in ['19', '15']:
                rid = f['rooms'].get(f'd{day}_{slot}', '')
                if rid:
                    room_ids[f'd{day}_{slot}'] = rid

        # Build product_name from mapping
        mapping = funnel_mapping.get(fnum)
        if mapping:
            prod, contr, var = mapping
            prod_id = product_ids[prod]
            contr_id = contractor_ids[contr]
            composed_name = f"{prod} {contr} {var}".strip() if var else f"{prod} {contr}"
        else:
            prod_id = product_ids[list(product_ids.keys())[0]]
            contr_id = contractor_ids[list(contractor_ids.keys())[0]]
            composed_name = f['name']
            var = ''

        bh_cond = bothelp_conditions.get(fnum, '')

        cur = conn.execute("""
            INSERT INTO funnels(
                num, source_id, product_id, contractor_id, variant,
                product_name, landing_url, start_date,
                block_name, sheet_name, tag_19_raw, tag_15_raw, reg_tags_raw,
                dash_sales_url, dash_pereliv_url,
                regi_total_url, regi_15_url, regi_19_url, regi_notime_url,
                predspisok_url, room_ids_json, bothelp_condition
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            fnum, source_ids[f['source']], prod_id, contr_id, var,
            composed_name, landing, sd,
            block_name, sheet_name, tag_19_raw, tag_15_raw, reg_raw,
            f['dash_sales'], f['dash_pereliv'],
            regi_total, f['regi_15'], f['regi_19'], f['regi_notime'],
            '', json.dumps(room_ids), bh_cond,
        ))
        funnel_id = cur.lastrowid

        # --- Insert funnel_tags ---
        for tag_type, raw in [('time_19', tag_19_raw), ('time_15', tag_15_raw), ('reg', reg_raw)]:
            tag_names = parse_tag_string(raw)
            for pos, tname in enumerate(tag_names):
                if tname in tag_ids:
                    conn.execute(
                        "INSERT OR IGNORE INTO funnel_tags(funnel_id, tag_id, tag_type, position) VALUES(?,?,?,?)",
                        (funnel_id, tag_ids[tname], tag_type, pos)
                    )
                    ftag_count += 1

        # --- Insert funnel_days (only days with actual content) ---
        do = day_overrides.get(fnum, {})
        for slot in ['19', '15']:
            days_data = b.get(f'days_{slot}', {})
            for day_num in range(1, 6):
                # Day override takes priority over extracted data
                override = do.get((slot, day_num))
                data = override if override else days_data.get(day_num)
                if not data:
                    continue
                room_f1 = f['rooms'].get(f'd{day_num}_{slot}', '')
                conn.execute("""
                    INSERT INTO funnel_days(
                        funnel_id, time_slot, day_num, room_id_f1,
                        gc_room, web_room, replay_url, web_replay,
                        sales_page, sales_note, tariffs, oto, bonuses,
                        mission, mission_type, meditation, dojim_note
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    funnel_id, slot, day_num, room_f1,
                    data.get('gc_room', ''), data.get('web_room', ''),
                    data.get('replay', ''), data.get('web_replay', ''),
                    data.get('sales_page', '') or data.get('meditation', ''),
                    data.get('sales_note', '') or data.get('dojim_note', ''),
                    data.get('tariffs', ''), data.get('oto', ''), data.get('bonuses', ''),
                    data.get('mission_gc', '') or data.get('mission_web', ''),
                    data.get('mission_type', ''),
                    data.get('meditation', ''), data.get('dojim_note', ''),
                ))
                days_count += 1

        # --- Insert salebot_configs ---
        so = salebot_overrides.get(fnum, {})
        for slot, sb_key in [('19', 'sb_19'), ('15', 'sb_15')]:
            sb_override = so.get(slot)
            if sb_override:
                cond = sb_override.get('condition', '')
                calc = sb_override.get('calculator', '')
            else:
                sb = f.get(sb_key, {})
                cond = sb.get('condition', '')
                calc = format_sb_calculator(sb.get('calculator', ''))
            if cond or calc:
                conn.execute(
                    "INSERT INTO salebot_configs(funnel_id, time_slot, condition, calculator) VALUES(?,?,?,?)",
                    (funnel_id, slot, cond, calc)
                )
                salebot_count += 1

    conn.commit()

    print(f"\n=== Database created: {db_path} ===")
    print(f"  sources:        {len(source_ids)}")
    print(f"  products:       {len(product_ids)}")
    print(f"  contractors:    {len(contractor_ids)}")
    print(f"  tags:           {len(tag_ids)}")
    print(f"  funnels:        {len(funnels)}")
    print(f"  funnel_tags:    {ftag_count}")
    print(f"  funnel_days:    {days_count}")
    print(f"  salebot_configs:{salebot_count}")

    return conn


# ============================================================
# 6. VERIFICATION QUERIES
# ============================================================

def verify(conn):
    print("\n=== Verification ===\n")

    # Counts
    for table in ['sources', 'products', 'contractors', 'tags', 'funnels', 'funnel_tags', 'funnel_days', 'salebot_configs']:
        cnt = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {table}: {cnt} rows")

    # Sample: funnels with tag "ДБО"
    print("\n-- Воронки с тегом 'ДБО': --")
    rows = conn.execute("""
        SELECT f.num, f.product_name, ft.tag_type
        FROM funnels f
        JOIN funnel_tags ft ON f.id = ft.funnel_id
        JOIN tags t ON ft.tag_id = t.id
        WHERE t.name = 'ДБО'
        ORDER BY f.num, ft.tag_type
    """).fetchall()
    for r in rows:
        print(f"  #{r[0]} {r[1]} ({r[2]})")

    # Sample: tags for funnel #11
    print("\n-- Теги воронки #11: --")
    rows = conn.execute("""
        SELECT ft.tag_type, GROUP_CONCAT(t.name, ', ') as tags
        FROM funnel_tags ft
        JOIN tags t ON ft.tag_id = t.id
        JOIN funnels f ON ft.funnel_id = f.id
        WHERE f.num = 11
        GROUP BY ft.tag_type
        ORDER BY ft.tag_type
    """).fetchall()
    for r in rows:
        print(f"  {r[0]}: {r[1]}")

    # Sample: funnels started after 2025-06-01
    print("\n-- Воронки, стартовавшие после 2025-06-01: --")
    rows = conn.execute("""
        SELECT num, product_name, start_date
        FROM funnels
        WHERE start_date > '2025-06-01'
        ORDER BY start_date
    """).fetchall()
    for r in rows:
        print(f"  #{r[0]} {r[1]} — {r[2]}")

    conn.close()


if __name__ == '__main__':
    conn = populate(DB_PATH)
    verify(conn)
