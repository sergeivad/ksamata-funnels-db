#!/usr/bin/env python3
"""
Генерация сводной таблицы автоворонок из SQLite базы данных.
Вход:  ksamata_funnels.db
Выход: Сводная_таблица_автоворонок.xlsx
"""

import sqlite3
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'ksamata_funnels.db')
OUT_PATH = os.path.join(BASE_DIR, 'Сводная_таблица_автоворонок.xlsx')

# ============================================================
# 1. LOAD DATA FROM DB
# ============================================================

def load_all(db_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # Funnels with source, product, contractor names
    funnels = conn.execute("""
        SELECT f.*, s.name as source_name,
               p.name as product_code, c.name as contractor_name
        FROM funnels f
        JOIN sources s ON f.source_id = s.id
        JOIN products p ON f.product_id = p.id
        JOIN contractors c ON f.contractor_id = c.id
        ORDER BY f.num
    """).fetchall()

    result = []
    for f in funnels:
        fid = f['id']
        fnum = f['num']

        # Tags grouped by type, ordered by position
        tags = {}
        for tag_type in ('time_19', 'time_15', 'reg'):
            rows = conn.execute("""
                SELECT t.name FROM funnel_tags ft
                JOIN tags t ON ft.tag_id = t.id
                WHERE ft.funnel_id = ? AND ft.tag_type = ?
                ORDER BY ft.position
            """, (fid, tag_type)).fetchall()
            tags[tag_type] = [r['name'] for r in rows]

        # Use raw tag strings from funnels table (preserves original formatting)
        tag_19_str = f['tag_19_raw'] or ''
        tag_15_str = f['tag_15_raw'] or ''
        reg_str = f['reg_tags_raw'] or ''

        # Days
        days = {}
        day_rows = conn.execute("""
            SELECT * FROM funnel_days
            WHERE funnel_id = ?
            ORDER BY time_slot DESC, day_num
        """, (fid,)).fetchall()
        for d in day_rows:
            key = (d['time_slot'], d['day_num'])
            days[key] = dict(d)

        # Product durations — attach duration_minutes to each day entry
        pid = f['product_id']
        dur_rows = conn.execute("""
            SELECT day_num, duration_minutes
            FROM product_durations
            WHERE product_id = ?
        """, (pid,)).fetchall()
        dur_map = {r['day_num']: r['duration_minutes'] for r in dur_rows}
        for (slot, day_num), day_data in days.items():
            day_data['duration_minutes'] = dur_map.get(day_num)

        # Salebot
        salebot = {}
        sb_rows = conn.execute("""
            SELECT * FROM salebot_configs
            WHERE funnel_id = ?
        """, (fid,)).fetchall()
        for sb in sb_rows:
            salebot[sb['time_slot']] = dict(sb)

        # Format start_date back to DD.MM.YYYY
        sd = f['start_date'] or ''
        if sd and '-' in sd:
            parts = sd.split('-')
            if len(parts) == 3:
                sd = f"{parts[2]}.{parts[1]}.{parts[0]}"

        result.append({
            'num': fnum,
            'source': f['source_name'],
            'name': f['product_name'],
            'product_code': f['product_code'],
            'contractor': f['contractor_name'],
            'variant': f['variant'] or '',
            'landing': f['landing_url'] or '',
            'start_date': sd,
            'block_name': f['block_name'] or '',
            'sheet': f['sheet_name'] or '',
            'tag_19': tag_19_str,
            'tag_15': tag_15_str,
            'reg_tags': reg_str,
            'dash_sales': f['dash_sales_url'] or '',
            'dash_pereliv': f['dash_pereliv_url'] or '',
            'predspisok': f['predspisok_url'] or '',
            'regi_total': f['regi_total_url'] or '',
            'regi_15': f['regi_15_url'] or '',
            'regi_19': f['regi_19_url'] or '',
            'regi_notime': f['regi_notime_url'] or '',
            'days': days,
            'salebot': salebot,
            'bothelp_condition': f['bothelp_condition'] or '',
            'room_ids': json.loads(f['room_ids_json'] or '{}'),
        })

    conn.close()
    return result


# ============================================================
# 2. BUILD EXCEL
# ============================================================

def format_sb(sb_data):
    if not sb_data:
        return ''
    cond = sb_data.get('condition', '')
    calc = sb_data.get('calculator', '')
    parts = []
    if cond:
        parts.append(f"condition: {cond}")
    if calc:
        parts.append(calc)
    return '\n'.join(parts) if parts else ''


def build_excel(funnels_data, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводная автоворонки"

    hdr_fill = PatternFill(start_color="2F5496", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    meta_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    col_hdr_fill = PatternFill(start_color="D6E4F0", fill_type="solid")
    fill_19 = PatternFill(start_color="E2EFDA", fill_type="solid")
    fill_15 = PatternFill(start_color="FCE4D6", fill_type="solid")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    wrap = Alignment(wrap_text=True, vertical='top')
    NC = 14

    col_headers = [
        'Время', 'День',
        'Комната GC (аналит.)', 'Комната Web (вебинар)',
        'Длительность',
        'Повтор', 'Продажная', 'Примечание',
        'Тарифы ГК', 'OTO', 'Бонусы / Программа',
        'GetCourse / Salebot',
        'Condition / Калькулятор',
        'BotHelp condition',
    ]

    def wmeta(r, label, val):
        if not val:
            return r
        ws.cell(r, 1, label).font = Font(bold=True, size=9)
        ws.cell(r, 1).fill = meta_fill
        ws.cell(r, 1).border = border
        ws.cell(r, 2, val).alignment = wrap
        ws.cell(r, 2).border = border
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
        return r + 1

    def fmt_duration(minutes):
        if not minutes:
            return ''
        h = minutes // 60
        m = minutes % 60
        return f"⏱ {h}ч {m}м"

    def wday(r, time, day, data, fill):
        vals = [
            time, f'{day} день',
            data.get('gc_room', ''), data.get('web_room', ''),
            fmt_duration(data.get('duration_minutes')),
            data.get('replay_url', '') or data.get('web_replay', ''),
            data.get('sales_page', '') or data.get('meditation', ''),
            data.get('sales_note', '') or data.get('dojim_note', ''),
            data.get('tariffs', ''), data.get('oto', ''), data.get('bonuses', ''),
            data.get('mission', ''),
            '',
            '',
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v if v else '')
            cell.fill = fill
            cell.border = border
            cell.alignment = wrap

    cur = 1
    for f in funnels_data:
        fnum = f['num']

        # Header row
        cell = ws.cell(cur, 1, f"#{fnum}  {f['name']}")
        cell.font = hdr_font
        cell.fill = hdr_fill
        for c in range(1, NC + 1):
            ws.cell(cur, c).fill = hdr_fill
            ws.cell(cur, c).border = border
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=NC)
        cur += 1

        # Meta rows
        cur = wmeta(cur, 'Продукт:', f['product_code'])
        cur = wmeta(cur, 'Подрядчик:', f['contractor'])
        if f['variant']:
            cur = wmeta(cur, 'Вариант:', f['variant'])
        cur = wmeta(cur, 'Тег 19:00:', f['tag_19'])
        cur = wmeta(cur, 'Тег 15:00:', f['tag_15'])
        cur = wmeta(cur, 'Теги регистрации:', f['reg_tags'])
        cur = wmeta(cur, 'Лист:', f['sheet'])
        cur = wmeta(cur, 'Посадочная:', f['landing'])
        cur = wmeta(cur, 'Дата старта:', f['start_date'])
        cur = wmeta(cur, 'Дашборд продаж:', f['dash_sales'])
        cur = wmeta(cur, 'Дашборд перелива:', f['dash_pereliv'])
        cur = wmeta(cur, 'Предсписок:', f['predspisok'] if f['predspisok'] else 'Нет предсписка')
        cur = wmeta(cur, 'Реги общ ГК:', f['regi_total'])
        cur = wmeta(cur, 'Реги на 15:00:', f['regi_15'])
        cur = wmeta(cur, 'Реги на 19:00:', f['regi_19'])
        cur = wmeta(cur, 'Реги без выбора:', f['regi_notime'])

        # Column headers
        for c, h in enumerate(col_headers, 1):
            cell = ws.cell(cur, c, h)
            cell.font = Font(bold=True, size=9)
            cell.fill = col_hdr_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cur += 1

        # Salebot text
        sb_19_text = format_sb(f['salebot'].get('19'))
        sb_15_text = format_sb(f['salebot'].get('15'))

        # Days 19:00
        start_19 = cur
        for day in range(1, 6):
            data = f['days'].get(('19', day))
            if data:
                wday(cur, '19:00', day, data, fill_19)
                cur += 1
        end_19 = cur

        # Days 15:00
        start_15 = cur
        for day in range(1, 6):
            data = f['days'].get(('15', day))
            if data:
                wday(cur, '15:00', day, data, fill_15)
                cur += 1
        end_15 = cur

        # Merge salebot column (col 13)
        SB_COL = 13
        if end_19 > start_19:
            if end_19 - start_19 > 1:
                ws.merge_cells(start_row=start_19, start_column=SB_COL, end_row=end_19 - 1, end_column=SB_COL)
            cell = ws.cell(start_19, SB_COL, sb_19_text)
            cell.alignment = wrap
            cell.border = border
            cell.fill = fill_19

        if end_15 > start_15:
            if end_15 - start_15 > 1:
                ws.merge_cells(start_row=start_15, start_column=SB_COL, end_row=end_15 - 1, end_column=SB_COL)
            cell = ws.cell(start_15, SB_COL, sb_15_text)
            cell.alignment = wrap
            cell.border = border
            cell.fill = fill_15

        # Merge BotHelp condition column (col 14) — spans all days
        BH_COL = 14
        bh_text = f.get('bothelp_condition', '')
        total_start = start_19 if end_19 > start_19 else start_15
        total_end = end_15 if end_15 > start_15 else end_19
        if total_end > total_start:
            if total_end - total_start > 1:
                ws.merge_cells(start_row=total_start, start_column=BH_COL, end_row=total_end - 1, end_column=BH_COL)
            cell = ws.cell(total_start, BH_COL, bh_text)
            cell.alignment = wrap
            cell.border = border

        cur += 1  # blank row between funnels

    # Column widths
    widths = [16, 10, 38, 38, 14, 30, 38, 28, 35, 30, 30, 40, 45, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"Saved: {len(funnels_data)} funnels, {cur} rows → {out_path}")


if __name__ == '__main__':
    data = load_all(DB_PATH)
    build_excel(data, OUT_PATH)
